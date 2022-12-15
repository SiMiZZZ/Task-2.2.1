import math
import re
import csv
import datetime
from statistics import mean
import matplotlib
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from jinja2 import Environment, FileSystemLoader
import pdfkit


class Vacancy():
    def __init__(self):
        self.name = str()
        self.salary_from = str()
        self.salary_to = str()
        self.salary_currency = str()
        self.area_name = str()
        self.published_at = str()

    def get_ru_salary(self):
        self.currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
        }
        return (int(float(self.salary_from)) + int(float(self.salary_to))) / 2 * self.currency_to_rub[
                    self.salary_currency]



class DataSet():
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = []

    def csv_reader(self, file_name):
        empty = False
        vacancies = []
        first_line = []
        with open(file_name, "r", encoding='utf_8_sig') as csv_file:
            reader = csv.reader(csv_file)
            for index, row in enumerate(reader):
                if index == 0:
                    first_line = row
                    quanity = len(first_line)
                else:
                    if self.check_list(row, quanity):
                        vacancies.append(row)
        if first_line == []:
            print("Пустой файл")
            empty = True
        elif vacancies == []:
            print("Нет данных")
            empty = True
        return first_line, vacancies, empty

    def check_list(self, non_checked_list, quanity):
        if len(non_checked_list) == quanity and ('' not in non_checked_list):
            return True
        return False

    def clear_list(self, value):
        value = re.sub(r'\<[^>]*\>', '', value)
        return value

    def csv_filer(self, reader, list_naming):
        vacancies = []
        for vacancie in list_naming:
            clear_naming = {}
            for index, skill in enumerate(reader):
                clear_naming[skill] = " ".join(self.clear_list(vacancie[index]).split())
            vacancies.append(clear_naming)
        return vacancies


class InputCorrect():
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession_name = input("Введите название профессии: ")

    def get_key(self, dict, value):
        for k, v in dict.items():
            if v == value:
                return k

    def __check_skills(self, skills, vacancy):
            for b in skills:
                if b not in vacancy:
                    return False
            return True

class Statistics():
    def __init__(self, vacancies, profession_name):
        self.vacancies = vacancies
        self.profession_name = profession_name
        self.suitable_cities = []
        self.share_of_cities = self.make_share_of_cities()
        self.salary_by_years = self.make_salary_by_years()
        self.quantity_by_years = self.make_quantity_by_years()
        self.salary_by_profession = self.make_salary_by_profession()
        self.quantity_by_profession = self.make_quantity_by_profession()
        self.salary_by_cities = self.make_salary_by_sities()

    def make_salary_by_years(self):
        salary_by_years = {}
        for vacancie in self.vacancies:
            vacancie_year = int(datetime.datetime.strptime(vacancie.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y"))
            if vacancie_year not in salary_by_years.keys():
                salary_by_years[vacancie_year] = []
            salary_by_years[vacancie_year].append(vacancie.get_ru_salary())
        for year in salary_by_years.keys():
            salary_by_years[year] = int(mean(salary_by_years[year]))
        salary_by_years = dict(sorted(salary_by_years.items(), key=lambda x: x[0]))
        return salary_by_years

    def make_quantity_by_years(self):
        quantity_by_years = {}
        for vacancie in self.vacancies:
            vacancie_year = int(datetime.datetime.strptime(vacancie.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y"))
            if vacancie_year not in quantity_by_years.keys():
                quantity_by_years[vacancie_year] = 0
            quantity_by_years[vacancie_year] += 1
        quantity_by_years = dict(sorted(quantity_by_years.items(), key=lambda x: x[0]))
        return quantity_by_years

    def make_salary_by_profession(self):
        salary_by_years = {}
        for vacancie in self.vacancies:
            if self.profession_name not in vacancie.name:
                continue
            vacancie_year = int(datetime.datetime.strptime(vacancie.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y"))
            if vacancie_year not in salary_by_years.keys():
                salary_by_years[vacancie_year] = []
            salary_by_years[vacancie_year].append(vacancie.get_ru_salary())
        for year in salary_by_years.keys():
            salary_by_years[year] = int(mean(salary_by_years[year]))
        salary_by_years = dict(sorted(salary_by_years.items(), key=lambda x: x[0]))
        if len(salary_by_years.keys()) == 0:
            salary_by_years[2022] = 0
        return salary_by_years

    def make_quantity_by_profession(self):
        quantity_by_years = {}
        for vacancie in self.vacancies:
            if self.profession_name not in vacancie.name:
                continue
            vacancie_year = int(datetime.datetime.strptime(vacancie.published_at, "%Y-%m-%dT%H:%M:%S%z").strftime("%Y"))
            if vacancie_year not in quantity_by_years.keys():
                quantity_by_years[vacancie_year] = 0
            quantity_by_years[vacancie_year] += 1
        quantity_by_years = dict(sorted(quantity_by_years.items(), key=lambda x: x[0]))
        if len(quantity_by_years.keys()) == 0:
            quantity_by_years[2022] = 0
        return quantity_by_years

    def make_salary_by_sities(self):
        salary_by_cities = {}
        for vacancie in self.vacancies:
            if vacancie.area_name not in self.suitable_cities:
                continue
            vacancie_city = vacancie.area_name
            if vacancie_city not in salary_by_cities.keys():
                salary_by_cities[vacancie_city] = []
            salary_by_cities[vacancie_city].append(vacancie.get_ru_salary())
        for area_name in salary_by_cities.keys():
            salary_by_cities[area_name] = int(mean(salary_by_cities[area_name]))
        salary_by_cities = sorted(salary_by_cities.items(), key=lambda x: x[1], reverse=True)
        salary_by_cities = dict(salary_by_cities[:min(10,len(salary_by_cities))])
        return salary_by_cities

    def make_share_of_cities(self):
        vacancies_quantity = len(self.vacancies)
        share_of_cities = {}
        pop_names= []
        for vacancie in self.vacancies:
            vacancie_city = vacancie.area_name
            if vacancie_city not in share_of_cities.keys():
                share_of_cities[vacancie_city] = 0
            share_of_cities[vacancie_city] += 1

        for area_name in share_of_cities.keys():
            share_of_cities[area_name] = round(share_of_cities[area_name]/vacancies_quantity,4)
            if share_of_cities[area_name]<0.01:
                pop_names.append(area_name)
            else:
                self.suitable_cities.append(area_name)

        same_vacancies = 0
        for a in pop_names:
            same_vacancies += share_of_cities[a]
            share_of_cities.pop(a)
        share_of_cities = sorted(share_of_cities.items(), key=lambda x: x[1], reverse=True)
        share_of_cities = dict(share_of_cities[:min(10,len(share_of_cities))])
        share_of_cities["Другие"] = same_vacancies
        return share_of_cities


class Plot:
    def __init__(self, statistic):
        self.statistic = statistic
        self.fig, self.axs = plt.subplots(2, 2, figsize=(20,9))
        self.create_salary_diagramm()
        self.create_years_diagramm()
        self.create_salary_by_cities_diagramm()
        self.create_share_of_cities_diagramm()
        plt.savefig("graph.png")
        plt.show()

    def create_salary_diagramm(self):
        horizontal_labels = list(statistic.salary_by_years.keys())
        horizontal_by_year = list(map(lambda x: x+0.2, list(statistic.salary_by_years.keys())))
        vertical_by_year = list(statistic.salary_by_years.values())

        horizontal_by_profession = list(map(lambda x: x-0.2, list(statistic.salary_by_profession.keys())))
        vertical_by_profession = list(statistic.salary_by_profession.values())

        self.axs[0, 0].bar(horizontal_by_year, vertical_by_year, label="средняя з/п", width=0.4)
        self.axs[0, 0].bar(horizontal_by_profession, vertical_by_profession, label="з/п {}".format(statistic.profession_name), width=0.4)
        self.axs[0, 0].set_xticks(ticks=horizontal_labels, labels=horizontal_labels, rotation=90, fontsize=8)
        plt.tick_params(labelsize=8)
        self.axs[0, 0].title.set_text("Уровень зарплат по годам")
        self.axs[0, 0].legend(fontsize=8, loc='upper left')
        self.axs[0, 0].grid(axis='y')

    def create_years_diagramm(self):
        horizontal_labels = list(statistic.quantity_by_years.keys())
        horisontal_by_year = list(map(lambda x: x+0.2, list(statistic.quantity_by_years.keys())))
        vertical_by_year = list(statistic.quantity_by_years.values())

        horisontal_by_profession = list(map(lambda x: x - 0.2, list(statistic.quantity_by_years.keys())))
        vertical_by_profession = list(statistic.quantity_by_profession.values())

        self.axs[0, 1].bar(horisontal_by_year, vertical_by_year, label="Количество вакансий", width=0.4)
        self.axs[0, 1].bar(horisontal_by_profession, vertical_by_profession, label="Количество вакансий {}".format(statistic.profession_name), width=0.4)
        self.axs[0, 1].set_xticks(ticks=horizontal_labels, labels=horizontal_labels, rotation=90, fontsize=8)
        self.axs[0, 1].title.set_text("Количество вакансий по годам")
        self.axs[0, 1].legend(fontsize=8, loc='upper left')
        self.axs[0, 1].grid(axis='y')

    def create_salary_by_cities_diagramm(self):
        horisontal = list(statistic.salary_by_cities.keys())
        vertical = list(statistic.salary_by_cities.values())
        horisontal = list(map(lambda x: x.replace("-", "-\n").replace(" ", "\n"), horisontal))

        self.axs[1, 0].barh(horisontal, vertical)
        self.axs[1, 0].grid(axis="x")
        self.axs[1, 0].title.set_text("Уровень зарплат по городам")

    def create_share_of_cities_diagramm(self):
        cities = list(statistic.share_of_cities.keys())
        shares = list(statistic.share_of_cities.values())

        self.axs[1, 1].title.set_text("Доля вакансий по городам")
        self.axs[1,1].pie(shares, labels=cities, textprops={'fontsize': 6})

class XlTable():
    def __init__(self):
        self.wb = Workbook()
        del self.wb['Sheet']
        self.sheet_by_years = self.wb.create_sheet("Статистика по годам")
        self.sheet_by_cities = self.wb.create_sheet("Статистика по городам")
        self.thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

    def set_cell_width(self, sheet):
        column_widths = []
        for row in sheet.rows:
            for i, cell in enumerate(row):
                cellv = str(cell.value)
                if len(column_widths) > i:
                    if len(cellv) > column_widths[i]:
                        column_widths[i] = len(cellv)
                else:
                    column_widths += [len(cellv)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            sheet.column_dimensions[get_column_letter(i)].width = column_width +3

    def make_sheet_by_cities(self, statistic):
        headers = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        self.sheet_by_cities.append(headers)
        headers_cells = [self.sheet_by_cities["A1"],self.sheet_by_cities["B1"],
                         self.sheet_by_cities["D1"],self.sheet_by_cities["E1"]]
        for cell in headers_cells:
            cell.border = thin_border
            cell.font = Font(bold=True)

        self.sheet_by_cities.append(headers)
        city_index = 0
        value_index = 0
        city_values = statistic.salary_by_cities
        for row in self.sheet_by_cities.iter_rows(max_row=len(city_values)+1, min_row=2, min_col=1, max_col=2):
            for cell in row:
                if cell.column == 1:
                    cell.value = list(statistic.salary_by_cities.keys())[city_index]
                    city_index+=1
                else:
                    cell.value = list(statistic.salary_by_cities.values())[value_index]
                    value_index+=1
                cell.border = self.thin_border

        city_index = 0
        value_index = 0
        city_values = statistic.share_of_cities
        for row in self.sheet_by_cities.iter_rows(max_row=len(city_values)+1, min_row=2, min_col=4, max_col=5):
            for cell in row:
                if cell.column == 4:
                    cell.value = list(statistic.share_of_cities.keys())[city_index]
                    city_index+=1
                else:
                    cell.number_format = "0.00%"
                    cell.value = list(statistic.share_of_cities.values())[value_index]
                    value_index+=1
                cell.border = self.thin_border


        self.set_cell_width(self.sheet_by_cities)

        self.wb.save("report.xlsx")

    def make_sheet_by_years(self, statistic):
        headers = ["Год", "Средняя зарплата", "Средняя зарплата - {}".format(statistic.profession_name),
                   "Количество вакансий", "Количество вакансий - {}".format(statistic.profession_name)]

        self.sheet_by_years.append(headers)
        for year in statistic.salary_by_years.keys():
            row = [year]
            mean_salary = statistic.salary_by_years[year]
            mean_profession_salary = statistic.salary_by_profession[year]
            vacancies_quantity = statistic.quantity_by_years[year]
            vacancies_profession_quantity = statistic.quantity_by_profession[year]
            row = row + [mean_salary, mean_profession_salary, vacancies_quantity, vacancies_profession_quantity]
            self.sheet_by_years.append(row)

        for row in self.sheet_by_years.iter_rows(max_col=5, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left")

        for row in self.sheet_by_years.rows:
            for cell in row:
                cell.border = self.thin_border

        self.set_cell_width(self.sheet_by_years)
        self.wb.save("report.xlsx")


def generate_pdf(prof, statistic):
    headers1 = ["Год", "Средняя зарплата", f"Средняя зарплата - {prof}", "Количество вакансий",
                    f"Количество вакансий - {prof}"]
    headers2 = ["Город", "Уровень зарплат", "Город", "Доля вакансий"]

    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template("pdf_template.html")

    pdf_template = template.render({'profession': prof,
                                        "headers1": headers1,
                                        "headers2": headers2,
                                        "salary_by_years": statistic.salary_by_years,
                                        "vacancies_by_years": statistic.salary_by_profession,
                                        "vacancies_salary_by_years": statistic.quantity_by_years,
                                        "vacancies_counts_by_years": statistic.quantity_by_profession,
                                        "salary_by_cities": statistic.salary_by_cities,
                                        "vacs_by_cities": statistic.share_of_cities
                                        })

    config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
    pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})

def set_class_values(data):
    vacancies = []
    for dic in data:
        vacancy = Vacancy()
        for value in dic.items():
            setattr(vacancy, value[0], value[1])
        vacancies.append(vacancy)
    return vacancies

choice = input('''1. Вывод табличных данных
2. Формирование графиков и отчетов
''')
input_rows = InputCorrect()
data_set = DataSet(input_rows.file_name)
f_line, vacancies, empty = data_set.csv_reader(input_rows.file_name)
if not empty:
    data = data_set.csv_filer(f_line, vacancies)
    data_set.vacancies_objects = set_class_values(data)
    statistic = Statistics(data_set.vacancies_objects, input_rows.profession_name)
    if choice == "2":
        pl = Plot(statistic)
        generate_pdf(statistic.profession_name, statistic)
    else:
        print("{}: {}".format("Динамика уровня зарплат по годам", statistic.salary_by_years))
        print("{}: {}".format("Динамика количества вакансий по годам", statistic.quantity_by_years))
        print("{}: {}".format("Динамика уровня зарплат по годам для выбранной профессии", statistic.salary_by_profession))
        print("{}: {}".format("Динамика количества вакансий по годам для выбранной профессии",
                              statistic.quantity_by_profession))
        print("{}: {}".format("Уровень зарплат по городам (в порядке убывания)", statistic.salary_by_cities))
        print("{}: {}".format("Доля вакансий по городам (в порядке убывания)", statistic.share_of_cities))


